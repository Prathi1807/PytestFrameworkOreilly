import openpyxl
class HomePageData:

    test_Homepagedata= [{'Firstname':'Rahul','lastname':'Shetty','gender':'Male'}, {'Firstname':'Prathipha','lastname':'anbu','gender':'Female'}]

    @staticmethod
    def getTestaData(test_case_name): #self para is nt required for static methods
        dict={}
        book = openpyxl.load_workbook("C:\\Users\\hp\\OneDrive\\Documents\\Python_Practice\\python_Demo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == "Testcase2":
                for j in range(2, sheet.max_column + 1):
                    # dict["lastname"]="shetty"
                    dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [dict]
